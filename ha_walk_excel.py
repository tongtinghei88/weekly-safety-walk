from __future__ import annotations

import io
import os
import re
import zipfile
from copy import copy, deepcopy
from datetime import datetime, timedelta
from pathlib import Path

from lxml import etree as ET
from openpyxl.cell.cell import MergedCell
from openpyxl.cell.rich_text import CellRichText, InlineFont, TextBlock
from openpyxl import load_workbook
from openpyxl.styles import Border, Font, Side
from openpyxl.drawing.image import Image as XLImage
from PIL import Image

from gmail_ha_actions import fetch_ha_details


ROOT = Path(__file__).resolve().parent
WALK_DIR = Path(os.environ.get("HA_WALK_DIR", str(Path("G:/") / "\u6211\u7684\u96f2\u7aef\u786c\u789f" / "ACEW Safety" / "Element 6 - Inspection Programme" / "Site Safety and Environment Walk")))
DEFAULT_OUTPUT_DIR = Path(os.environ.get("HA_WALK_EXCEL_OUTPUT_DIR", str(WALK_DIR)))
WEEKLY_TEMPLATE = WALK_DIR / "Site Safety and Environment Walk No.79(05-05-2026).xlsx"
BIWEEKLY_TEMPLATE = WALK_DIR / "Site Safety and Environment Walk No.80(12-05-2026).xlsx"
NS_PKG_REL = "http://schemas.openxmlformats.org/package/2006/relationships"
ITEMS_PER_RECTIFICATION_SHEET = 4
CHECK_MARK = "✓"
LEGACY_CHECK_MARK = "ü"
CHECK_MARK_FONT = "Times New Roman"
THIN_SIDE = Side(style="thin")


def parse_date(date: str) -> datetime:
    return datetime.strptime(date, "%Y%m%d")


def date_label(date: str) -> str:
    return parse_date(date).strftime("%d-%m-%Y")


def find_report_no(date: str) -> int:
    label = date_label(date)
    for path in WALK_DIR.glob(f"Site Safety and Environment Walk No.*({label}).xlsx"):
        match = re.search(r"No\.(\d+)", path.name)
        if match:
            return int(match.group(1))

    dated: list[tuple[datetime, int]] = []
    for path in WALK_DIR.glob("Site Safety and Environment Walk No.*.xlsx"):
        if "completed" in path.name.lower() or "sample" in path.name.lower():
            continue
        match = re.search(r"No\.(\d+)\s*\((\d{2}-\d{2}-\d{4})\)", path.name)
        if match:
            dated.append((datetime.strptime(match.group(2), "%d-%m-%Y"), int(match.group(1))))
    before = [(dt, number) for dt, number in dated if dt <= parse_date(date)]
    if before:
        return sorted(before)[-1][1] + 1
    return max((number for _, number in dated), default=0) + 1


def walk_type_for_file_date(file_date: datetime) -> str:
    week_number = ((file_date.day - 1) // 7) + 1
    return "weekly" if week_number % 2 == 1 else "biweekly"


def dated_walk_files() -> list[tuple[datetime, int, Path]]:
    files: list[tuple[datetime, int, Path]] = []
    for path in WALK_DIR.glob("Site Safety and Environment Walk No.*.xlsx"):
        if "completed" in path.name.lower() or "sample" in path.name.lower():
            continue
        match = re.search(r"No\.(\d+)\s*\((\d{2}-\d{2}-\d{4})\)", path.name)
        if match:
            files.append((datetime.strptime(match.group(2), "%d-%m-%Y"), int(match.group(1)), path))
    return files


def template_for_type(walk_type: str, date: str | None = None, issue_count: int | None = None) -> Path:
    target_date = parse_date(date) if date else None
    slot_count = min(issue_count, ITEMS_PER_RECTIFICATION_SHEET) if issue_count is not None else None
    preferred = BIWEEKLY_TEMPLATE if walk_type == "biweekly" else WEEKLY_TEMPLATE
    preferred_match = re.search(r"\((\d{2}-\d{2}-\d{4})\)", preferred.name)
    preferred_date = datetime.strptime(preferred_match.group(1), "%d-%m-%Y") if preferred_match else None
    if preferred.exists() and (target_date is None or (preferred_date is not None and preferred_date < target_date)):
        if slot_count is None or photo_slot_count(preferred) == slot_count:
            return preferred

    exact_candidates: list[tuple[datetime, int, Path]] = []
    fallback_candidates: list[tuple[datetime, int, Path]] = []
    for file_date, report_no, path in dated_walk_files():
        if target_date is not None and file_date >= target_date:
            continue
        if walk_type_for_file_date(file_date) != walk_type:
            continue
        if slot_count is not None and photo_slot_count(path) == slot_count:
            exact_candidates.append((file_date, report_no, path))
        fallback_candidates.append((file_date, report_no, path))

    if exact_candidates:
        return sorted(exact_candidates, key=lambda item: (item[0], item[1]))[-1][2]

    if slot_count is not None:
        any_type_exact = [
            (file_date, report_no, path)
            for file_date, report_no, path in dated_walk_files()
            if (target_date is None or file_date < target_date) and photo_slot_count(path) == slot_count
        ]
        if any_type_exact:
            return sorted(any_type_exact, key=lambda item: (item[0], item[1]))[-1][2]

    candidates = fallback_candidates
    if not candidates:
        candidates = [
            (file_date, report_no, path)
            for file_date, report_no, path in dated_walk_files()
            if walk_type_for_file_date(file_date) == walk_type
            and (target_date is None or file_date != target_date)
        ]
    if not candidates:
        raise FileNotFoundError(f"Missing Excel template for {walk_type}: {preferred}")
    return sorted(candidates, key=lambda item: (item[0], item[1]))[-1][2]


def photo_slot_count(path: Path) -> int:
    try:
        with zipfile.ZipFile(path, "r") as zf:
            before, after = drawing_photo_slots(zf)
        return min(len(before), len(after))
    except (OSError, zipfile.BadZipFile, ET.XMLSyntaxError):
        return 0


def legacy_template_for_type(walk_type: str, date: str | None = None) -> Path:
    target_date = parse_date(date) if date else None
    preferred = BIWEEKLY_TEMPLATE if walk_type == "biweekly" else WEEKLY_TEMPLATE
    preferred_match = re.search(r"\((\d{2}-\d{2}-\d{4})\)", preferred.name)
    preferred_date = datetime.strptime(preferred_match.group(1), "%d-%m-%Y") if preferred_match else None
    if preferred.exists() and (target_date is None or preferred_date != target_date):
        return preferred

    candidates = [
        (file_date, report_no, path)
        for file_date, report_no, path in dated_walk_files()
        if walk_type_for_file_date(file_date) == walk_type
        and (target_date is None or file_date != target_date)
    ]
    if not candidates:
        raise FileNotFoundError(f"Missing Excel template for {walk_type}: {preferred}")
    return sorted(candidates, key=lambda item: (item[0], item[1]))[-1][2]


def template_for_type_label(walk_type: str) -> str:
    return "Bi-Weekly" if walk_type == "biweekly" else "Weekly"


def drawing_targets(zf: zipfile.ZipFile) -> dict[str, str]:
    rel_names = sorted(name for name in zf.namelist() if name.startswith("xl/drawings/_rels/drawing") and name.endswith(".rels"))
    if not rel_names:
        return {}
    rels = ET.fromstring(zf.read(rel_names[0]))
    targets: dict[str, str] = {}
    for rel in rels.findall(f"{{{NS_PKG_REL}}}Relationship"):
        rid = rel.attrib["Id"]
        targets[rid] = rel.attrib["Target"].replace("../", "xl/")
    return targets


def drawing_photo_slots(zf: zipfile.ZipFile) -> tuple[list[str], list[str]]:
    rel_names = sorted(name for name in zf.namelist() if name.startswith("xl/drawings/_rels/drawing") and name.endswith(".rels"))
    if not rel_names:
        return [], []
    rel_name = rel_names[0]
    drawing_name = rel_name.replace("xl/drawings/_rels/", "xl/drawings/").replace(".rels", "")
    rels = ET.fromstring(zf.read(rel_name))
    relmap = {rel.attrib["Id"]: rel.attrib["Target"].replace("../", "xl/") for rel in rels}
    root = ET.fromstring(zf.read(drawing_name))
    ns = {
        "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
        "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
        "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    }
    anchors: list[tuple[int, int, str]] = []
    for anchor in root.xpath(".//xdr:twoCellAnchor|.//xdr:oneCellAnchor", namespaces=ns):
        marker = anchor.find("xdr:from", ns)
        blip = anchor.find(".//a:blip", ns)
        if marker is None or blip is None:
            continue
        rid = blip.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed")
        if not rid or rid not in relmap:
            continue
        row = int(marker.find("xdr:row", ns).text)
        col = int(marker.find("xdr:col", ns).text)
        anchors.append((row, col, relmap[rid]))

    before: list[str] = []
    after: list[str] = []
    for _, group in groupby_rows(anchors):
        row_slots = sorted(group, key=lambda item: item[1])
        if row_slots:
            before.append(row_slots[0][2])
        if len(row_slots) > 1:
            after.append(row_slots[-1][2])
    return before, after


def groupby_rows(anchors: list[tuple[int, int, str]]) -> list[tuple[int, list[tuple[int, int, str]]]]:
    groups: list[tuple[int, list[tuple[int, int, str]]]] = []
    for anchor in sorted(anchors):
        row = anchor[0]
        if groups and row - groups[-1][0] <= 2:
            groups[-1][1].append(anchor)
        else:
            groups.append((row, [anchor]))
    return groups


def blank_image() -> Image.Image:
    return Image.new("RGB", (900, 650), "white")


def load_photo(path: Path) -> Image.Image:
    return Image.open(path).copy()


def location_base(location: str) -> str:
    return re.sub(r"\s*\(Section\s+\d+\)\s*", "", location, flags=re.IGNORECASE).strip()


def rectification_image_slots(ws: Any) -> tuple[list[Any], list[Any]]:
    anchors: list[tuple[int, int, Any]] = []
    for image in getattr(ws, "_images", []):
        marker = image.anchor._from
        anchors.append((int(marker.row), int(marker.col), image))

    before: list[Any] = []
    after: list[Any] = []
    for _, group in groupby_rows([(row, col, image) for row, col, image in anchors]):
        row_slots = sorted(group, key=lambda item: item[1])
        if row_slots:
            before.append(row_slots[0][2])
        if len(row_slots) > 1:
            after.append(row_slots[-1][2])
    return before, after


def replacement_image(source: Path | None, template_image: Any) -> XLImage:
    if source:
        image = XLImage(str(source))
    else:
        buffer = io.BytesIO()
        blank_image().save(buffer, format="PNG")
        buffer.seek(0)
        image = XLImage(buffer)
        image._ha_buffer = buffer
    image.anchor = deepcopy(template_image.anchor)
    image.width = template_image.width
    image.height = template_image.height
    return image


def set_cover_title(cover: Any, walk_type: str) -> None:
    base_font = cover["B2"].font
    normal = InlineFont(rFont=base_font.name, sz=base_font.sz, b=base_font.bold)
    struck = InlineFont(rFont=base_font.name, sz=base_font.sz, b=base_font.bold, strike=True)
    if walk_type == "biweekly":
        cover["B2"] = CellRichText(
            TextBlock(struck, "Weekly"),
            TextBlock(normal, " / Biweekly Safety & Environmental Walk & Follow-up Actions "),
        )
    else:
        cover["B2"] = CellRichText(
            TextBlock(normal, "Weekly / "),
            TextBlock(struck, "Biweekly"),
            TextBlock(normal, " Safety & Environmental Walk & Follow-up Actions "),
        )


SAFETY_CHECKLIST_RULES: tuple[tuple[int, tuple[str, ...], tuple[str, ...]], ...] = (
    (9, ("form 1", "weekly inspection", "statutory inspection"), ("lifting",)),
    (10, ("colour code", "color code"), ("lifting gear", "lifting appliance", "sling")),
    (19, ("webbing sling", "lifting sling", "broken sling", "lifting gear", "lifting appliance"), ()),
    (27, ("concrete pipe", "concrete pipes", "gully", "wedge", "wedges"), ()),
    (42, ("warning sign", "warning notice", "warning notices", "caution sign", "notice"), ("fencing", "barrier", "cover", "opening")),
    (43, ("opening", "manhole", "covering", "covered", "cover"), ("opening", "manhole", "properly covered")),
    (46, ("proper access", "excavation access", "access for excavation", "safe access"), ("excavation",)),
    (66, ("ip67", "waterproof socket", "socket", "cable", "electric wire", "extension wire"), ()),
    (67, ("damaged socket", "damaged cable", "patent damage", "broken socket"), ()),
    (68, ("wireless grinder", "cordless", "coreless", "portable electrical tool"), ()),
    (73, ("live part", "electric wire", "welding wire", "welding cable"), ("welding",)),
    (78, ("fire extinguisher", "extinguisher"), ("gas", "welding", "flame cutting", "hot work")),
    (79, ("pressure gauge", "gauge", "regulator"), ("gas", "welding", "flame cutting", "oxygen", "acetylene")),
    (113, ("flammable label", "label", "labeling", "labelling"), ("flammable", "chemical", "dangerous goods")),
    (117, ("material stacking", "stacking", "stacked", "placed on ground", "stored on ground"), ()),
    (118, ("formwork", "material storage", "materials", "tidied", "properly stored", "housekeeping"), ()),
    (119, ("debris", "rubbish", "waste", "removed from site", "dispose"), ()),
    (120, ("projecting nail", "nail", "sharp object", "starter bar", "protruding bar"), ()),
    (121, ("ladder", "working platform", "passageway", "access route", "working area"), ()),
    (122, ("laid on ground", "trip hazard", "obstruction", "obstructed"), ("wire", "cable", "hose", "passageway", "access")),
    (124, ("water ponding", "ponding", "stagnant water", "mosquito", "slippery"), ()),
    (125, ("site boundary", "fenced off", "entrance", "public access"), ("fencing", "barrier", "boundary")),
    (131, ("helmet", "chin strap", "safety helmet"), ()),
    (132, ("eye protection", "goggle", "goggles", "face shield"), ()),
    (136, ("anti-collision", "anti collision", "collision rod", "shoes", "gloves", "heat shelter", "others"), ()),
    (139, ("fencing", "barrier", "between the access", "material storage area", "lantern"), ()),
    (140, ("exit notice", "exit sign", "signboard", "access notice", "working access"), ()),
    (142, ("fire extinguisher", "extinguisher"), ("generator", "excavator", "plant", "machine", "equipment")),
)


def safety_checklist_row(issue: str) -> int:
    text = " ".join(issue.lower().split())
    if any(term in text for term in ("laid on ground", "trip hazard", "obstruction", "obstructed")) and any(
        term in text for term in ("wire", "cable", "hose", "passageway", "access")
    ):
        return 122
    if any(term in text for term in ("exit notice", "exit sign", "signboard", "working access")):
        return 140
    if ("fire extinguisher" in text or "extinguisher" in text) and any(
        term in text for term in ("gas", "welding", "flame cutting", "hot work")
    ):
        return 78
    if "fencing" in text and any(term in text for term in ("between the access", "material storage", "storage area")):
        return 139
    for row, any_terms, required_terms in SAFETY_CHECKLIST_RULES:
        if any(term in text for term in any_terms) and (
            not required_terms or any(term in text for term in required_terms)
        ):
            return row
    if "fire extinguisher" in text or "extinguisher" in text:
        return 142
    if "warning" in text or "notice" in text:
        return 42
    if "fencing" in text or "barrier" in text:
        return 139
    if "material" in text or "storage" in text or "housekeeping" in text:
        return 118
    return 136


def apply_safety_checklist(safety: Any, issues: list[str]) -> None:
    def set_check(cell: Any) -> None:
        size = cell.font.sz or 11
        cell.value = CHECK_MARK
        cell.font = Font(name=CHECK_MARK_FONT, sz=size)

    for row in range(8, safety.max_row + 1):
        for col in (6, 7, 8, 9):
            cell = safety.cell(row, col)
            if cell.value in (CHECK_MARK, LEGACY_CHECK_MARK):
                set_check(cell)

    for row in range(8, safety.max_row + 1):
        item_no = safety.cell(row, 2).value
        if item_no in (None, ""):
            continue
        if safety.cell(row, 8).value not in (None, "") or safety.cell(row, 10).value not in (None, ""):
            safety.cell(row, 8).value = None
            safety.cell(row, 10).value = None
            if safety.cell(row, 9).value in (None, ""):
                set_check(safety.cell(row, 6))

    row_to_photos: dict[int, list[int]] = {}
    for index, issue in enumerate(issues, start=1):
        row_to_photos.setdefault(safety_checklist_row(issue), []).append(index)

    for row, photo_numbers in row_to_photos.items():
        safety.cell(row, 6).value = None
        safety.cell(row, 7).value = None
        set_check(safety.cell(row, 8))
        safety.cell(row, 9).value = None
        safety.cell(row, 10).value = ", ".join(f"Photo No.{number}" for number in photo_numbers)


def prepare_cover_issue_rows(cover: Any, issue_count: int) -> list[int]:
    if issue_count > ITEMS_PER_RECTIFICATION_SHEET:
        extra_rows = issue_count - ITEMS_PER_RECTIFICATION_SHEET
        insert_at = 23
        shifted_ranges = [merged for merged in list(cover.merged_cells.ranges) if merged.min_row >= insert_at]
        for merged in shifted_ranges:
            cover.unmerge_cells(str(merged))
        cover.insert_rows(insert_at, extra_rows)
        for merged in shifted_ranges:
            cover.merge_cells(
                start_row=merged.min_row + extra_rows,
                start_column=merged.min_col,
                end_row=merged.max_row + extra_rows,
                end_column=merged.max_col,
            )
        for offset in range(extra_rows):
            row = insert_at + offset
            cover.row_dimensions[row].height = cover.row_dimensions[22].height
            for col in range(2, 15):
                source = cover.cell(22, col)
                target = cover.cell(row, col)
                target._style = copy(source._style)
                target.number_format = source.number_format
                target.alignment = copy(source.alignment)
                target.font = copy(source.font)
                target.fill = copy(source.fill)
                target.border = copy(source.border)
            cover.merge_cells(start_row=row, start_column=3, end_row=row, end_column=9)
            cover.merge_cells(start_row=row, start_column=10, end_row=row, end_column=11)
            cover.merge_cells(start_row=row, start_column=12, end_row=row, end_column=14)
    return list(range(19, 19 + max(issue_count, ITEMS_PER_RECTIFICATION_SHEET)))


def signed_date_cells(issue_count: int) -> tuple[str, str, str, str]:
    if issue_count <= 3:
        return "C28", "H28", "C29", "H29"
    extra_rows = max(0, issue_count - ITEMS_PER_RECTIFICATION_SHEET)
    return f"C{29 + extra_rows}", f"H{29 + extra_rows}", "C28", "H28"


def apply_cover_signatures(cover: Any, walk_type: str, inspect_date: datetime, issue_count: int) -> None:
    def set_value(cell_ref: str, value: Any) -> None:
        if not isinstance(cover[cell_ref], MergedCell):
            cover[cell_ref] = value

    primary_date_1, primary_date_2, clear_date_1, clear_date_2 = signed_date_cells(issue_count)
    cover[primary_date_1] = inspect_date
    cover[primary_date_2] = inspect_date
    cover[clear_date_1] = None
    cover[clear_date_2] = None

    if walk_type == "weekly":
        for cell_ref in ("L30", "L31", "M31", "N31", "O31", "L33", "M33"):
            set_value(cell_ref, None)
        for row in range(30, 34):
            for col in range(13, 16):
                cover.cell(row, col).border = Border()
        return

    set_value("L30", "Signed : ")
    set_value("L31", "(")
    set_value("M31", "Patrick, P. T. KO\nCE/T243")
    set_value("L33", "Date : ")
    for col in range(13, 15):
        cover.cell(30, col).border = copy(cover.cell(30, col).border)
        cover.cell(30, col).border = Border(bottom=THIN_SIDE)
        cover.cell(31, col).border = Border(top=THIN_SIDE)
        cover.cell(33, col).border = Border(bottom=THIN_SIDE)


def apply_anti_mosquito_signatures(anti_mosquito: Any, walk_type: str) -> None:
    def set_value(cell_ref: str, value: Any) -> None:
        if not isinstance(anti_mosquito[cell_ref], MergedCell):
            anti_mosquito[cell_ref] = value

    if walk_type == "weekly":
        for cell_ref in ("H28", "I29", "J29", "L29", "H31", "J31"):
            set_value(cell_ref, None)
        for row in range(28, 32):
            for col in range(10, 12):
                anti_mosquito.cell(row, col).border = Border()
        return

    set_value("H28", "Signed : ")
    set_value("I29", "(")
    set_value("J29", "Patrick, P. T. KO\nCE/T243")
    set_value("L29", ")")
    set_value("H31", "Date : ")
    for col in range(10, 12):
        anti_mosquito.cell(29, col).border = Border(top=THIN_SIDE)
        anti_mosquito.cell(31, col).border = Border(bottom=THIN_SIDE)


def fill_rectification_sheet(
    sheet: Any,
    page_number: int,
    issues: list[str],
    actions: list[str],
    locations: list[str],
    before_photos: list[Path],
    after_photos: list[Path],
    template_before_slots: list[Any],
    template_after_slots: list[Any],
) -> None:
    rect_photo_rows = [19, 40, 61, 82]
    rect_rows = [20, 41, 62, 83]
    rect_issue_rows = [21, 42, 63, 84]
    start_index = page_number * ITEMS_PER_RECTIFICATION_SHEET

    for local_index, row in enumerate(rect_rows):
        global_index = start_index + local_index
        photo_row = rect_photo_rows[local_index]
        issue_row = rect_issue_rows[local_index]
        if global_index >= len(issues):
            for ref in (f"D{photo_row}", f"L{photo_row}", f"D{row}", f"L{row}", f"D{issue_row}", f"L{issue_row}"):
                sheet[ref] = None
            continue
        location = location_base(locations[global_index]) if global_index < len(locations) else ""
        sheet[f"D{photo_row}"] = global_index + 1
        sheet[f"L{photo_row}"] = global_index + 1
        sheet[f"D{row}"] = location
        sheet[f"L{row}"] = location
        sheet[f"D{issue_row}"] = issues[global_index]
        sheet[f"L{issue_row}"] = actions[global_index]

    sheet._images = []
    for local_index, slot in enumerate(template_before_slots):
        global_index = start_index + local_index
        source = before_photos[global_index] if global_index < len(issues) and global_index < len(before_photos) else None
        sheet.add_image(replacement_image(source, slot))
    for local_index, slot in enumerate(template_after_slots):
        global_index = start_index + local_index
        source = after_photos[global_index] if global_index < len(issues) else None
        sheet.add_image(replacement_image(source, slot))


def create_walk_excel(
    date: str,
    walk_type: str,
    after_photos: list[Path],
    before_photos: list[Path] | None = None,
    out_dir: Path | None = None,
) -> Path:
    details = fetch_ha_details(date)
    issues = list(details["issues"])
    actions = list(details["actions"])
    locations = list(details["locations"])
    if len(after_photos) < len(issues):
        raise RuntimeError("Not enough after photos for Excel generation.")
    before_photos = before_photos or []

    template = template_for_type(walk_type, date, len(issues))
    if not template.exists():
        raise FileNotFoundError(f"Missing Excel template: {template}")

    report_no = find_report_no(date)
    out_dir = out_dir or DEFAULT_OUTPUT_DIR
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"Site Safety and Environment Walk No.{report_no}({date_label(date)}).xlsx"

    inspect_date = parse_date(date)
    due_date = inspect_date + timedelta(days=2)
    wb = load_workbook(template, rich_text=True)
    cover = wb["Cover"]
    rectification = wb["Rectification"]
    safety = wb["Checklist for safety"] if "Checklist for safety" in wb.sheetnames else None
    anti_mosquito = wb["Anti-Mosquito"] if "Anti-Mosquito" in wb.sheetnames else None

    cover["B3"] = f"Report No.: {report_no}"
    set_cover_title(cover, walk_type)
    cover["M6"] = inspect_date
    cover["M7"] = details.get("time") or "09:30 A.M."
    cover["E7"] = details.get("location_summary") or ""
    cover_issue_rows = prepare_cover_issue_rows(cover, len(issues))
    for index, row in enumerate(cover_issue_rows):
        if index < len(issues):
            cover[f"B{row}"] = index + 1
            cover[f"C{row}"] = issues[index]
            cover[f"J{row}"] = "HCECL"
            cover[f"L{row}"] = due_date
        else:
            cover[f"B{row}"] = None
            cover[f"C{row}"] = None
            cover[f"J{row}"] = None
            cover[f"L{row}"] = None
    apply_cover_signatures(cover, walk_type, inspect_date, len(issues))

    if anti_mosquito is not None:
        anti_mosquito["K5"] = inspect_date
        anti_mosquito["E6"] = details.get("location_summary") or ""
        apply_anti_mosquito_signatures(anti_mosquito, walk_type)
    if safety is not None:
        apply_safety_checklist(safety, issues)

    before_slots, after_slots = rectification_image_slots(rectification)
    page_count = max(1, (len(issues) + ITEMS_PER_RECTIFICATION_SHEET - 1) // ITEMS_PER_RECTIFICATION_SHEET)
    rectification_sheets = [rectification]
    for page_number in range(1, page_count):
        sheet = wb.copy_worksheet(rectification)
        sheet.title = f"Rectification {page_number + 1}"
        rectification_sheets.append(sheet)
    for page_number, sheet in enumerate(rectification_sheets):
        fill_rectification_sheet(
            sheet,
            page_number,
            issues,
            actions,
            locations,
            before_photos,
            after_photos,
            before_slots,
            after_slots,
        )

    temp_out_path = out_path.with_name(f"~{out_path.stem}.tmp.xlsx")
    try:
        wb.save(temp_out_path)
        os.replace(temp_out_path, out_path)
    except PermissionError as exc:
        temp_out_path.unlink(missing_ok=True)
        raise PermissionError(f"Cannot overwrite Excel file. Please close it first: {out_path}") from exc
    return out_path


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 3:
        raise SystemExit("Usage: python ha_walk_excel.py YYYYMMDD weekly|biweekly")
    job_dir = ROOT / "telegram_photos" / sys.argv[1]
    photos = sorted(job_dir.glob("photo_*.jpg"))
    print(create_walk_excel(sys.argv[1], sys.argv[2], photos))
